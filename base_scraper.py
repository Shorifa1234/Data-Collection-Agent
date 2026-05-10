"""
base_scraper.py
---------------
AVS Agent — Shared utilities used by every vendor scraper.

Key design: columns are DYNAMIC.
- The scraper collects every field it finds on the product page.
- ExcelWriter buffers all rows and determines the full column set at save-time.
- Column order = CORE_FIRST + tracker studio_columns + extra discovered fields.

Provides:
  - PlaywrightBrowser  : async context manager wrapping Playwright
  - ExcelWriter        : dynamic column, row-buffered Excel writer
  - build_column_order : merge studio_columns + discovered keys into ordered list
  - parse_dimensions   : "W 25" x D 12" x H 22.5"" → {Width, Depth, Height, …}
  - parse_spec_block   : "DESIGNER: X | COLLECTION: Y | …" → dict
  - clean_text         : normalise whitespace
  - safe_float         : tolerant float conversion
  - polite_delay / async_polite_delay
"""

from __future__ import annotations

import asyncio
import re
import time
import random
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Known column ordering  (defines preferred left-to-right sequence)
# ---------------------------------------------------------------------------
# Columns always written first (identity / pricing)
CORE_FIRST = [
    "Index", "Category", "Manufacturer",
    "Source URL", "Image URL",
    "Product Name", "Product Family Id",
    "Price", "SKU",
    "Description", "Weight",
    "Specifications", "Materials",
    "Dimensions", "Length", "Width", "Depth", "Diameter", "Height",
]

# Seating / furniture extras
SEATING_COLS = [
    "Seat Height", "Seat Depth", "Seat Length", "Arm Height",
]

# Structural / component extras
STRUCTURAL_COLS = [
    "Base", "Canopy", "Shade Details",
]

# Upholstery / finish extras
FINISH_COLS = [
    "COM", "COL", "Fabric", "FABRIC", "Finish", "FINISH",
    "Back", "Color", "Wood", "Footrest",
    "Upholstery", "UPHOLSTERY", "Timber",
    "Components",
]

# Lighting-specific extras
LIGHTING_COLS = [
    "Illumination", "Suspension", "Socket", "Wattage", "Size",
]

# Attribution / provenance
ATTRIBUTION_COLS = [
    "Designer", "Maker", "Collection",
    "Lead Time", "Origin", "Production", "Date",
    "Tariff Disclaimer",
]

# Always last
CORE_LAST = ["Tearsheet Link"]

# Complete preferred ordering (used by build_column_order)
PREFERRED_ORDER: list[str] = (
    CORE_FIRST
    + SEATING_COLS
    + STRUCTURAL_COLS
    + FINISH_COLS
    + LIGHTING_COLS
    + ATTRIBUTION_COLS
    + CORE_LAST
)
_PREFERRED_SET = {c.lower(): c for c in PREFERRED_ORDER}


def build_column_order(
    studio_columns: list[str],
    discovered_keys: set[str],
) -> list[str]:
    """
    Merge studio_columns (from tracker) with discovered_keys (from scraper)
    into an ordered column list.

    Priority:
      1. CORE_FIRST — always at the front
      2. studio_columns — tracker-defined columns, in tracker order
      3. PREFERRED_ORDER remainder — known columns in logical order
      4. Anything still unseen — appended alphabetically at the end
    """
    result: list[str] = []
    seen: set[str] = set()

    def add(col: str):
        c = col.strip()
        if c and c not in seen:
            result.append(c)
            seen.add(c)

    # 1. Always-first core
    for c in CORE_FIRST:
        add(c)

    # 2. studio_columns in order (skip ones already added)
    for c in studio_columns:
        add(c)

    # 3. Rest of preferred order
    for c in PREFERRED_ORDER:
        add(c)

    # 4. Discovered keys not yet placed — sorted alphabetically
    remaining = sorted(k for k in discovered_keys if k not in seen)
    for c in remaining:
        add(c)

    return result


# ---------------------------------------------------------------------------
# ExcelWriter  — fully dynamic, row-buffering
# ---------------------------------------------------------------------------
class ExcelWriter:
    """
    Buffers all product rows per sheet and writes them when save() is called.

    Columns are determined dynamically from every row written — no fixed schema.
    Column ordering uses build_column_order() so studio_columns from the tracker
    define preferred order, and any extra scraped fields are appended at the end.

    Usage:
        writer = ExcelWriter(output_path, vendor_name)
        writer.add_sheet("Nightstands", category_link, studio_columns=["SKU", ...])
        writer.write_row({"Product Name": "...", "SKU": "...", "Seat Height": "18"})
        writer.save()
    """

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    META_FONT   = Font(bold=True, size=10)

    def __init__(self, output_path: Path, vendor_name: str):
        self.output_path = output_path
        self.vendor_name = vendor_name
        # _sheets: ordered dict of sheet_name → sheet_data
        self._sheets: dict[str, dict] = {}
        self._current: str | None = None

    def add_sheet(
        self,
        category_name: str,
        category_link: str = "",
        studio_columns: list[str] | None = None,
    ):
        """Register a new category sheet. Rows are buffered until save()."""
        key = category_name[:31]   # Excel sheet name limit
        self._sheets[key] = {
            "display_name": category_name,
            "category_link": category_link,
            "studio_columns": studio_columns or [],
            "rows": [],          # list of dicts
            "all_keys": set(),   # union of all keys seen in rows
        }
        self._current = key

    # Mandatory fields — written even if empty so the column is always present
    MANDATORY_FIELDS = [
        "Index", "Category", "Manufacturer", "Source URL", "Image URL",
        "Product Name", "Product Family Id", "SKU",
    ]

    def write_row(self, data: dict[str, Any], category_name: str = ""):
        """
        Buffer one product row on the current sheet.

        Mandatory fields (Index, Category, Manufacturer, Source URL, Image URL,
        Product Name, Product Family Id, SKU) are always included in
        all_keys so their columns always appear even if some rows are missing them.

        Manufacturer is auto-populated from vendor_name if not already set.

        Dimension sub-fields (Width, Height, Depth, Diameter, Length) are
        auto-derived from the Dimensions string if they are not already present.
        """
        if self._current is None:
            raise RuntimeError("Call add_sheet() before write_row()")
        sheet = self._sheets[self._current]
        # Backward compat: rename legacy key "Source" → "Source URL"
        if "Source" in data and "Source URL" not in data:
            data = {("Source URL" if k == "Source" else k): v for k, v in data.items()}
        # Keep non-empty values only (None / "" / [] are excluded)
        row = {k: v for k, v in data.items() if v not in (None, "", [])}
        if category_name:
            row.setdefault("Category", category_name)
        # Auto-populate Manufacturer from vendor name
        row.setdefault("Manufacturer", self.vendor_name)
        # Auto-derive individual dimension fields from Dimensions string if missing
        if row.get("Dimensions"):
            parsed = parse_dimensions(str(row["Dimensions"]))
            for sub in ("Width", "Height", "Depth", "Diameter", "Length"):
                if parsed.get(sub) and sub not in row:
                    row[sub] = parsed[sub]
        sheet["rows"].append(row)
        sheet["all_keys"].update(row.keys())
        # Ensure mandatory columns always appear in the sheet
        sheet["all_keys"].update(self.MANDATORY_FIELDS)

    def save(self):
        """Write every buffered sheet to the output .xlsx file."""
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        for sheet_key, sheet_data in self._sheets.items():
            rows = sheet_data["rows"]
            if not rows:
                print(f"  [Excel] Skipping empty sheet: {sheet_key}")
                continue

            # Number the rows
            for i, row in enumerate(rows, 1):
                row["Index"] = i

            # Build column order from studio_columns + all discovered keys
            columns = build_column_order(
                sheet_data["studio_columns"],
                sheet_data["all_keys"],
            )

            # Drop columns that are entirely empty across all rows
            # (mandatory fields are always kept even if blank)
            mandatory_set = set(self.MANDATORY_FIELDS)
            columns = [
                col for col in columns
                if col in mandatory_set
                or any(row.get(col) not in (None, "", []) for row in rows)
            ]

            ws = wb.create_sheet(title=sheet_key)

            # ── meta rows ─────────────────────────────────────────────
            ws.cell(1, 1, "Brand ").font = self.META_FONT
            ws.cell(1, 2, self.vendor_name)
            ws.cell(2, 1, "Link").font = self.META_FONT
            if sheet_data["category_link"]:
                ws.cell(2, 2, sheet_data["category_link"])
            # row 3 = blank

            # ── header row (row 4) ────────────────────────────────────
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(4, col_idx, col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            ws.freeze_panes = "A5"
            ws.row_dimensions[4].height = 28

            # ── data rows (starting at row 5) ─────────────────────────
            for row_num, row_data in enumerate(rows, start=5):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    cell = ws.cell(row_num, col_idx, value)
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(col_name in ("Description", "Specifications")),
                    )

            # ── column widths ─────────────────────────────────────────
            _WIDTH_MAP = {
                "Index": 6, "Category": 18, "Manufacturer": 22,
                "Source URL": 42, "Image URL": 42,
                "Product Name": 38, "Product Family Id": 32,
                "Price": 10, "SKU": 15,
                "Description": 55, "Weight": 10,
                "Specifications": 65, "Materials": 22,
                "Dimensions": 28,
            }
            for col_idx, col_name in enumerate(columns, start=1):
                width = _WIDTH_MAP.get(col_name, 16)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            print(
                f"  [Excel] Sheet '{sheet_key}': "
                f"{len(rows)} rows x {len(columns)} cols"
            )

        if not wb.sheetnames:
            print(f"[Excel] WARNING: no data rows found — skipping save for {self.output_path}")
            return

        try:
            wb.save(self.output_path)
        except PermissionError:
            raise PermissionError(
                f"Cannot save '{self.output_path}' — the file is open in another program. "
                "Close it and re-run."
            )
        print(f"[Excel] Saved -> {self.output_path}")


# ---------------------------------------------------------------------------
# Dimension parser
# ---------------------------------------------------------------------------
_DIM_MAP = {
    "w": "Width", "d": "Depth", "h": "Height",
    "l": "Length",
    "dia": "Diameter", "diam": "Diameter", "diameter": "Diameter",
}

def parse_dimensions(raw: str) -> dict[str, str]:
    """
    Parse dimension strings in any of these formats:
      'W 25" x D 12" x H 22.5"'        ← label before number
      'Dia. 18" x H 12"'
      'L 48" x W 24" x H 30"'
      '16.00" L x 10.75" W x 5.00" H'  ← number before label (e.g. Hennepin Made)
      '30H x 18.5Dia'                   ← number directly followed by label (no quote)

    Returns:
      Dimensions  — cleaned original string with inch marks removed
      Width/Depth/Height/Length/Diameter — numeric string only (no " or in)
    """
    if not raw:
        return {}

    # Normalise inch symbols and quotes to plain "
    text = (raw
            .replace("\u2033", '"')   # DOUBLE PRIME \u2033
            .replace("\u201d", '"')   # RIGHT DOUBLE QUOTATION MARK "
            .replace("\u201c", '"')   # LEFT DOUBLE QUOTATION MARK "
            .replace("\u02ba", '"')   # MODIFIER LETTER DOUBLE PRIME \u02fa
            .replace("\u2032", "'")   # PRIME \u2032
            .replace("\u2019", '"'))  # RIGHT SINGLE QUOTATION MARK ' (used as inch on some sites)

    # Dimensions field: strip inch marks and "in" suffixes, keep numbers + separators
    dim_clean = re.sub(r'["\']+', '', text)                   # remove " and '
    dim_clean = re.sub(r'\bin\.?\b', '', dim_clean, flags=re.IGNORECASE)
    dim_clean = re.sub(r'\s+', ' ', dim_clean).strip()
    result: dict[str, str] = {"Dimensions": dim_clean}

    seen: set[str] = set()   # avoid overwriting first match with a later duplicate axis

    def _add(key_raw: str, raw_val: str) -> None:
        key = _DIM_MAP.get(key_raw.lower().rstrip("."))
        if key and key not in seen:
            seen.add(key)
            result[key] = _fraction_to_decimal(raw_val.strip())

    # Pattern A: label before number  e.g.  W 25"  /  Dia. 18"  /  H22.5
    _pat_a = re.compile(
        r'\b(dia(?:meter|m)?\.?|[wdhl])\s*[.:]?\s*([\d]+(?:[./][\d]+)?(?:\.\d+)?)\s*(?:"|in\.?)?',
        re.IGNORECASE,
    )
    for m in _pat_a.finditer(text):
        _add(m.group(1), m.group(2))

    # Pattern B: number before label  e.g.  16.00" W  /  36" dia  /  30H  /  18.5Dia
    _pat_b = re.compile(
        r'\b([\d]+(?:\.\d+)?)\s*(?:"|in\.?)?\s*(dia(?:meter|m)?\.?|[wdhl])\b',
        re.IGNORECASE,
    )
    for m in _pat_b.finditer(text):
        _add(m.group(2), m.group(1))   # groups are reversed vs pattern A

    return result


def _fraction_to_decimal(value: str) -> str:
    """Convert '1/2' → '0.5', '25' → '25', '22.5' → '22.5'."""
    if "/" in value:
        try:
            parts = value.split("/")
            return str(round(int(parts[0]) / int(parts[1]), 4))
        except (ValueError, ZeroDivisionError):
            pass
    return value


# ---------------------------------------------------------------------------
# Spec-block parser
# ---------------------------------------------------------------------------

# All possible spec key aliases → canonical output column name
_SPEC_KEY_MAP: dict[str, str] = {
    "designer":           "Designer",
    "maker":              "Maker",
    "collection":         "Collection",
    "lead time":          "Lead Time",
    "origin":             "Origin",
    "production":         "Production",
    "date":               "Date",
    "materials":          "Materials",
    "material":           "Materials",
    "dimensions":         "Dimensions",
    "weight":             "Weight",
    "finish":             "Finish",
    "finishes":           "Finish",
    "base":               "Base",
    "canopy":             "Canopy",
    "shade":              "Shade Details",
    "shade details":      "Shade Details",
    "illumination":       "Illumination",
    "suspension":         "Suspension",
    "socket":             "Socket",
    "wattage":            "Wattage",
    "seat height":        "Seat Height",
    "seat depth":         "Seat Depth",
    "seat length":        "Seat Length",
    "arm height":         "Arm Height",
    "com":                "COM",
    "col":                "COL",
    "fabric":             "Fabric",
    "upholstery":         "Upholstery",
    "color":              "Color",
    "colour":             "Color",
    "components":         "Components",
    "tariff disclaimer":  "Tariff Disclaimer",
    "size":               "Size",
    "footrest":           "Footrest",
    "back":               "Back",
    "wood":               "Wood",
    "timber":             "Timber",
}


def parse_spec_block(specs: str) -> dict[str, str]:
    """
    Parse a pipe-separated key: value specification block.
    Also handles newline-separated and colon-only formats.
    Returns a dict with canonical column-name keys.
    """
    if not specs:
        return {}
    result: dict[str, str] = {}

    # Split on pipe or newline
    for segment in re.split(r"[|\n]", specs):
        segment = segment.strip()
        if not segment or ":" not in segment:
            continue
        k, _, v = segment.partition(":")
        k_norm = k.strip().lower()
        v_norm = v.strip()
        if not v_norm:
            continue

        # Try exact match first, then startswith
        canonical = _SPEC_KEY_MAP.get(k_norm)
        if canonical is None:
            for alias, canon in _SPEC_KEY_MAP.items():
                if k_norm.startswith(alias):
                    canonical = canon
                    break
        if canonical is None:
            # Keep unknown keys as title-cased
            canonical = k.strip().title()

        result[canonical] = v_norm

    return result


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------
def clean_text(text: str | None) -> str:
    if not text:
        return ""
    text = str(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def sentence_case(text: str | None) -> str:
    """
    Convert text to sentence case: first letter capitalised, rest lowercased.
    Handles None/empty safely.

    Examples:
      "OSLO DINING CHAIR"          → "Oslo dining chair"
      "REEF TABLE LAMP - BRASS"    → "Reef table lamp - brass"
      "oak nightstand"             → "Oak nightstand"
    """
    if not text:
        return ""
    t = clean_text(text)
    return t[0].upper() + t[1:].lower() if t else ""


def safe_float(value: str | None) -> float | None:
    if not value:
        return None
    try:
        return float(re.sub(r"[^\d.]", "", str(value)))
    except (ValueError, TypeError):
        return None


def clean_price(raw: str | None) -> float | str | None:
    """
    Return Price as a plain number — no '$', 'USD', commas, or range notation.
    Takes the lower value if a range is given (e.g. '$1,200 – $1,800' → 1200.0).
    Returns None if no numeric value can be extracted.
    """
    if not raw:
        return None
    # Take the first value in a range
    first = re.split(r"[–—\-]", str(raw))[0]
    # Strip everything that isn't a digit or decimal point
    numeric = re.sub(r"[^\d.]", "", first.strip())
    return safe_float(numeric)


# ---------------------------------------------------------------------------
# SKU generation
# ---------------------------------------------------------------------------
def generate_sku(vendor_name: str, category_name: str, index: int) -> str:
    """
    Generate a SKU when the vendor does not provide one.

    Formula: first 3 letters of vendor (alpha only, uppercase)
           + first 3 letters of category (alpha only, uppercase)
           + product index number

    Examples:
      vendor="Visual Comfort", category="Chandeliers", index=45 → "VISCHA45"
      vendor="Made Goods",     category="Lounge Chairs", index=9 → "MADLOU9"
    """
    v = re.sub(r"[^A-Z]", "", vendor_name.upper())[:3]
    c = re.sub(r"[^A-Z]", "", category_name.upper())[:3]
    # Pad vendor/category parts to 3 chars if the name is short
    v = v.ljust(3, "X")[:3]
    c = c.ljust(3, "X")[:3]
    return f"{v}{c}{index}"


# ---------------------------------------------------------------------------
# Product Family Id extraction
# ---------------------------------------------------------------------------

# Variant suffixes that should be stripped to find the family name
_VARIANT_PATTERNS = [
    # Size labels after dash/comma: "- Small", ", Large"
    r"\s*[-\u2013,]\s*(x-?small|small|medium|large|x-?large|xxl?|xs|sm|md|lg)\b.*$",
    # Colour / finish after dash: "- Ivory", "- Brass Finish"
    r"\s*[-\u2013]\s*[a-z ]{2,25}(?:finish|fabric|upholstery|leather|velvet|linen|wool)\b.*$",
    # Trailing parenthetical: "(Set of 2)", "(Round)"
    r"\s*\([^)]{1,40}\)\s*$",
    # Trailing size with unit: "48W" or bed sizes
    r'\s*[-\u2013,]?\s*\d+\s*[wWdDhHxX]+\s*$',
    r"\s*[-\u2013,]?\s*(king|queen|twin|full|cal\.?\s*king)\b.*$",
]
_VARIANT_RE = [re.compile(p, re.IGNORECASE) for p in _VARIANT_PATTERNS]


def extract_family_id(product_name: str) -> str:
    """
    Derive Product Family Id from Product Name.

    Product Family Id is the common base name shared across variants of the
    same product. Variant suffixes (size, colour, finish, dimensions) are
    stripped. If no variant suffix is detected the full name is returned.

    Examples:
      "OSLO DINING CHAIR - GREY FABRIC"  → "OSLO DINING CHAIR"
      "ATLAS SOFA (SET OF 2)"            → "ATLAS SOFA"
      "REEF TABLE LAMP - BRASS FINISH"   → "REEF TABLE LAMP"
      "OAK NIGHTSTAND"                   → "OAK NIGHTSTAND"   (unchanged)
    """
    name = product_name.strip()
    for pattern in _VARIANT_RE:
        cleaned = pattern.sub("", name).strip().rstrip(",-–").strip()
        if cleaned and cleaned != name:
            name = cleaned
            break   # apply only the first matching rule
    return name if name else product_name


# ---------------------------------------------------------------------------
# Rate-limit helpers
# ---------------------------------------------------------------------------
def polite_delay(min_sec: float = 1.0, max_sec: float = 3.0):
    time.sleep(random.uniform(min_sec, max_sec))


async def async_polite_delay(min_sec: float = 0.8, max_sec: float = 2.5):
    await asyncio.sleep(random.uniform(min_sec, max_sec))


# ---------------------------------------------------------------------------
# Playwright browser helper
# ---------------------------------------------------------------------------
class PlaywrightBrowser:
    """
    Async context manager that provides a Playwright Page with anti-detection headers.

    Usage:
        async with PlaywrightBrowser(headless=True) as page:
            await page.goto("https://example.com")
    """

    def __init__(self, headless: bool = True, slow_mo: int = 0):
        self.headless = headless
        self.slow_mo  = slow_mo
        self._playwright = None
        self._browser    = None
        self._context    = None
        self.page        = None

    # Resource types that don't affect HTML/JSON-LD extraction
    _BLOCK_TYPES = {"image", "media", "font", "other"}

    async def __aenter__(self):
        from playwright.async_api import async_playwright
        self._playwright = await async_playwright().start()
        self._browser = await self._playwright.chromium.launch(
            headless=self.headless,
            slow_mo=self.slow_mo,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--disable-extensions",
                "--disable-background-networking",
            ],
        )
        self._context = await self._browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
            locale="en-US",
        )
        await self._context.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        self.page = await self._context.new_page()
        # Block resource types that aren't needed for data extraction
        await self.page.route(
            "**/*",
            lambda route: (
                route.abort()
                if route.request.resource_type in self._BLOCK_TYPES
                else route.continue_()
            ),
        )
        return self.page

    async def __aexit__(self, *args):
        if self._context:
            await self._context.close()
        if self._browser:
            await self._browser.close()
        if self._playwright:
            await self._playwright.stop()
