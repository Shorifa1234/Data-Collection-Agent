"""
vendor_parser.py
----------------
Parses SD_Web Scraping - Status Tracker.xlsx to extract vendor info,
including per-category "Studio Column Names" defined in the tracker.

Usage:
    python vendor_parser.py "The Future Perfect"
    python vendor_parser.py --list          # list all vendor names

Each category now includes:
    "studio_columns": ["Product URL", "Image URL", "Product Name", ...]
These are the column names the tracker says should be captured for that category.
The scraper should collect ALL data from the site; studio_columns drive column order.
"""

import sys
import json
import argparse
from pathlib import Path
import openpyxl

TRACKER_PATH = Path(__file__).parent / "vendor sheet" / "SD_Web Scraping - Status Tracker.xlsx"

# Columns that appear in the tracker "Studio Column Names" row starting at col F (index 5)
STUDIO_COL_START_IDX = 6   # 0-based index of col G (col F is the "Studio Column Names" label)


def list_vendors(tracker_path: Path = TRACKER_PATH) -> list[str]:
    """Return all vendor sheet names (excludes meta-sheets)."""
    META_SHEETS = {
        "Main", "Furniture", "Lighting", "Wall Covering",
        "Template", "Templates", "Categories", "Attributes", "Potential Brands"
    }
    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    vendors = [s for s in wb.sheetnames if s not in META_SHEETS]
    wb.close()
    return vendors


def parse_vendor(vendor_name: str, tracker_path: Path = TRACKER_PATH) -> dict:
    """
    Parse one vendor sheet and return structured info:

    {
        "vendor_name": "The Future Perfect",
        "categories": [
            {
                "group": "Furniture",
                "name": "Nightstands",
                "links": ["https://..."],
                "studio_columns": ["Product URL", "Image URL", "Product Name", ...]
            },
            ...
        ]
    }

    - Categories with no links are omitted.
    - studio_columns comes from the "Studio Column Names" row in the tracker.
      These define the preferred column set for that category. The scraper may
      discover additional columns from the website — those are added dynamically.
    """
    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)

    sheet_names = wb.sheetnames
    match = None
    for s in sheet_names:
        if s.strip().lower() == vendor_name.strip().lower():
            match = s
            break

    if match is None:
        wb.close()
        raise ValueError(
            f"Vendor '{vendor_name}' not found in tracker.\n"
            f"Run with --list to see available vendors."
        )

    ws = wb[match]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # -----------------------------------------------------------------------
    # Tracker sheet layout — two variants are supported:
    #
    # Layout A (standard, e.g. The Future Perfect):
    #   col B (idx 1): product group label ("Furniture", "Lighting", …) — sparse
    #   col C (idx 2): row type → "Name ", "Category", "Link", "Link 2", etc.
    #   col D (idx 3): value for that row type
    #   col F+ (idx 6+): Studio Column Names (col F is the label, col G+ are values)
    #
    # Layout B (compact, e.g. Visual Comfort — no group column):
    #   col B (idx 1): row type → "Name", "Category", "Link", etc.
    #   col C (idx 2): value for that row type
    #   col E (idx 4): "Studio Column Name" label
    #   col F+ (idx 5+): studio column values
    #
    # Detection: if a "Category" or "Link" keyword appears in col B (not col C),
    # we switch to Layout B.
    # -----------------------------------------------------------------------

    # Detect layout by scanning the first few rows
    # Use only Category/Link keywords (not "name") to distinguish layouts
    _ROW_TYPE_KEYWORDS = {"category", "link", "link 2", "link 3"}
    layout_b = False
    for _r in rows[:10]:
        _b = str(_val(_r, 1)).strip().lower()
        _c = str(_val(_r, 2)).strip().lower()
        if _b in _ROW_TYPE_KEYWORDS and _c not in _ROW_TYPE_KEYWORDS:
            layout_b = True
            break

    if layout_b:
        # Layout B: type in col B, value in col C, studio cols from idx 5+
        TYPE_IDX, VAL_IDX, GROUP_IDX, STUDIO_START = 1, 2, None, 5
    else:
        # Layout A: group in col B, type in col C, value in col D, studio cols from idx 6+
        TYPE_IDX, VAL_IDX, GROUP_IDX, STUDIO_START = 2, 3, 1, STUDIO_COL_START_IDX

    vendor_display_name = vendor_name
    categories: list[dict] = []
    current_group = ""
    current_category: dict | None = None

    def flush_category():
        nonlocal current_category
        if current_category and current_category["links"]:
            categories.append(current_category)
        current_category = None

    for row in rows:
        row_type = str(_val(row, TYPE_IDX)).strip().lower()
        row_val  = _val(row, VAL_IDX)
        row_group = _val(row, GROUP_IDX) if GROUP_IDX is not None else ""

        # ── vendor name row ─────────────────────────────────────────────
        if row_type.startswith("name"):
            if row_val:
                vendor_display_name = str(row_val).strip()
            continue

        # ── product group label (Layout A only) ─────────────────────────
        if row_group and str(row_group).strip() and row_type == "category":
            current_group = str(row_group).strip()

        # ── "Category" row → new category block ────────────────────────
        if row_type == "category" and row_val:
            flush_category()

            studio_cols = []
            for idx in range(STUDIO_START, len(row)):
                v = row[idx]
                if v and str(v).strip() and str(v).strip() not in (" ", "Studio Column Name", "Studio Column Names"):
                    studio_cols.append(str(v).strip())

            current_category = {
                "group": current_group,
                "name": str(row_val).strip(),
                "links": [],
                "studio_columns": studio_cols,
            }
            continue

        # ── "Link" / "Link 2" / "Link 3" rows ──────────────────────────
        if row_type.startswith("link"):
            url = str(row_val).strip() if row_val else ""
            if url and url.startswith("http") and current_category is not None:
                current_category["links"].append(url)

            # Pick up studio_columns from Link row if Category row had none
            if current_category is not None and not current_category["studio_columns"]:
                studio_cols = []
                for idx in range(STUDIO_START, len(row)):
                    v = row[idx]
                    if v and str(v).strip() and str(v).strip() not in (" ", "Vendor Column Name", "Vendor Column Names"):
                        studio_cols.append(str(v).strip())
                if studio_cols:
                    current_category["studio_columns"] = studio_cols
            continue

        # ── Continuation link rows (no label in type col, but URL in value col) ──
        # Some trackers place extra links on plain rows below the "Link" row
        if current_category is not None and row_type == "":
            url = str(row_val).strip() if row_val else ""
            if url and url.startswith("http") and url not in current_category["links"]:
                current_category["links"].append(url)
            continue

    flush_category()

    return {
        "vendor_name": vendor_display_name,
        "categories": categories,
    }


def _val(row: tuple, idx: int):
    """Safe 0-based getter for a row tuple."""
    if idx < len(row):
        v = row[idx]
        return v if v is not None else ""
    return ""


def main():
    parser = argparse.ArgumentParser(description="Parse vendor data from SD tracker")
    parser.add_argument("vendor", nargs="?", help="Vendor name")
    parser.add_argument("--list", action="store_true", help="List all vendor names")
    parser.add_argument("--tracker", default=str(TRACKER_PATH), help="Path to tracker Excel")
    args = parser.parse_args()

    tracker = Path(args.tracker)

    if args.list:
        vendors = list_vendors(tracker)
        print("\n".join(vendors))
        return

    if not args.vendor:
        parser.print_help()
        sys.exit(1)

    try:
        data = parse_vendor(args.vendor, tracker)
        print(json.dumps(data, indent=2, ensure_ascii=False))
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
