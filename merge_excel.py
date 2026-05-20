"""
merge_excel.py — Merge two vendor Excel files into one.

Usage:
    python merge_excel.py <file1.xlsx> <file2.xlsx> [output.xlsx]

If output is omitted, saves as <vendor>_merged.xlsx in the same folder as file1.
Sheets from file1 come first, then any sheets from file2 that are not already in file1.
If a sheet exists in both files, file1 wins (kept as-is).
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def merge(file1: Path, file2: Path, output: Path):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    existing_sheets = {s.lower() for s in wb1.sheetnames}
    added = []

    for sheet_name in wb2.sheetnames:
        if sheet_name.lower() in existing_sheets:
            print(f"  SKIP  '{sheet_name}' — already in {file1.name}")
            continue

        src = wb2[sheet_name]
        dst = wb1.create_sheet(title=sheet_name)

        for row in src.iter_rows():
            for cell in row:
                new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                # Copy basic styling
                if cell.has_style:
                    new_cell.font      = Font(
                        bold=cell.font.bold,
                        color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else "000000",
                        size=cell.font.size,
                    )
                    new_cell.fill      = PatternFill(
                        fill_type=cell.fill.fill_type or None,
                        fgColor=cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.fgColor.type == "rgb" else "FFFFFF",
                    )
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                    )

        # Copy column widths
        for col_letter, col_dim in src.column_dimensions.items():
            dst.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights
        for row_num, row_dim in src.row_dimensions.items():
            dst.row_dimensions[row_num].height = row_dim.height

        # Copy freeze panes
        if src.freeze_panes:
            dst.freeze_panes = src.freeze_panes

        added.append(sheet_name)
        print(f"  ADDED '{sheet_name}' ({src.max_row - 4} rows)")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb1.save(output)
    print(f"\nSaved -> {output}")
    print(f"Sheets in output: {wb1.sheetnames}")


if __name__ == "__main__":
    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: python merge_excel.py <file1.xlsx> <file2.xlsx> [output.xlsx]")
        sys.exit(1)

    f1 = Path(args[0])
    f2 = Path(args[1])

    if len(args) >= 3:
        out = Path(args[2])
    else:
        out = f1.parent / (f1.stem + "_merged.xlsx")

    if not f1.exists():
        print(f"ERROR: {f1} not found"); sys.exit(1)
    if not f2.exists():
        print(f"ERROR: {f2} not found"); sys.exit(1)

    print(f"File 1 : {f1}  ({openpyxl.load_workbook(f1, read_only=True).sheetnames})")
    print(f"File 2 : {f2}  ({openpyxl.load_workbook(f2, read_only=True).sheetnames})")
    print(f"Output : {out}\n")

    merge(f1, f2, out)
